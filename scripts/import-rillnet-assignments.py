"""Build the server-side warehouse/PIC scope dataset from the approved workbook."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

import openpyxl


ZONE_ALIASES = {"Miên Bắc 4": "Miền Bắc 4"}


def identifier(value: object) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value or "").strip()


def identifiers(value: object) -> list[str]:
    if value in (None, "", "#N/A"):
        return []
    return sorted({part for part in re.split(r"[,;\s]+", identifier(value)) if part and part != "#N/A"})


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook")
    parser.add_argument("output")
    args = parser.parse_args()

    workbook = openpyxl.load_workbook(args.workbook, read_only=True, data_only=True)
    directory: dict[str, dict[str, str]] = {}
    for row in workbook["Danh bạ NV"].iter_rows(min_row=2, values_only=True):
        employee_id = identifier(row[0] if row else None)
        if employee_id:
            directory[employee_id] = {
                "employeeId": employee_id,
                "name": str(row[1] or "").strip() if len(row) > 1 else "",
                "title": str(row[3] or "").strip() if len(row) > 3 else "",
            }

    warehouses: dict[str, dict[str, object]] = {}
    duplicate_ids: set[str] = set()
    for row in workbook["Phân công"].iter_rows(min_row=2, values_only=True):
        warehouse_id = identifier(row[0] if row else None)
        if not warehouse_id:
            continue
        record = warehouses.get(warehouse_id)
        if record is None:
            record = {
                "warehouseId": warehouse_id,
                "warehouseName": str(row[1] or "").strip(),
                "warehouseType": str(row[2] or "").strip(),
                "province": str(row[3] or "").strip(),
                "zone": ZONE_ALIASES.get(str(row[4] or "").strip(), str(row[4] or "").strip()),
                "level1": [],
                "level2": [],
                "level3": [],
            }
            warehouses[warehouse_id] = record
        else:
            duplicate_ids.add(warehouse_id)
        for index, level in ((5, "level1"), (6, "level2"), (7, "level3")):
            record[level] = sorted(set(record[level]) | set(identifiers(row[index])))

    assigned_ids = {
        employee_id
        for warehouse in warehouses.values()
        for level in ("level1", "level2", "level3")
        for employee_id in warehouse[level]
    }
    people = [directory.get(employee_id, {"employeeId": employee_id, "name": "", "title": ""}) for employee_id in sorted(assigned_ids)]
    payload = {
        "schemaVersion": 1,
        "source": Path(args.workbook).name,
        "warehouses": sorted(warehouses.values(), key=lambda item: item["warehouseId"]),
        "people": people,
        "quality": {
            "duplicateWarehouseIdsMerged": sorted(duplicate_ids),
            "assignedPeopleMissingDirectory": sorted(assigned_ids - set(directory)),
        },
    }
    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")


if __name__ == "__main__":
    main()
